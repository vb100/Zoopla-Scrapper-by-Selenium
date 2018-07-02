""" Zoopla scraping project """

# Import libraries
import requests, re, os
import pandas as pd
from bs4 import BeautifulSoup

""" Generate the list of URLs : Start"""
def generateURLs(pages):
    listURLs = []
    base_url = "https://www.zoopla.co.uk/to-rent/property/london/west-drayton/?identifier=london%2Fwest-drayton&q=West%20Drayton%2C%20London&search_source=to-rent&radius=0&price_frequency=per_month&pn="
    for i in range(1, pages+1, 1):
        fullURL = base_url + str(i)
        listURLs.append(fullURL)
    return listURLs
""" Generate the list of URLs : End"""

""" Get Data : Start"""
properties = [] # Here all the data will be stored
def getData(listURLs):
    
    import re
    #from selenium import webdriver
    #options = webdriver.chrome.options.Options()
    #options.add_argument("--disable-extensions")
    
    featuresOfProperties = {}
    
    #chrome_path = r"C:\Users\user\Desktop\Python\JSscrapping\chromedriver.exe"
    #driver = webdriver.Chrome(chrome_path)
        
    def getFullDescription(objectURL):
            
        #print(objectURL)
        r = requests.get(objectURL)
        c = r.content
        soup = BeautifulSoup(c, "html.parser")
            
        try:
            description = soup.find("div", {"itemprop":"description"}).text
        except:
            description = ""
        cleaned = description.replace(",", " ")
        cleaned = cleaned.replace(":", " ")
        cleaned = cleaned.replace(".", " ")
        cleaned = cleaned.replace("W C", "WC")
        cleaned = cleaned.replace("/", " ")
        cleaned = cleaned.replace(")", ") ")
        cleaned = cleaned.replace("Description", "Description ")
        cleaned = cleaned.replace("Floor", "Floor ")
        cleaned = cleaned.replace("Floor", "Floor ")
        cleaned = cleaned.replace("Annexe", "")
        cleaned = cleaned.replace("Divided", " Divided")
        cleaned = cleaned.replace("  ", " ")
        return cleaned
        
    for page in range(1, len(listURLs), 1): #(1, len(listURLs), 1):
        print(page," : ", listURLs[page])
        r = requests.get(listURLs[page])
        c = r.content
        soup = BeautifulSoup(c, "html.parser")
            
        infoTable = soup.find("ul", {"class":"listing-results"})
        items = infoTable.find_all("div", {"class":"listing-results-wrapper"})
        
        #print("Len. = ", len(items))
            
        for item in range (0, len(items), 1):
            #print("Starting:", item)
            blockProperty = items[item]
                    
            """ get Address """
            address = blockProperty.find("a", {"class":"listing-results-address"}).text
                    
            """ Listed on """
            listedOn = blockProperty.find("p", {"class":"top-half"}).find("small").text.split("\n")[2].replace(" ", "")
            #print("listedOn:", listedOn)
            monthNum = " "
            if "Nov" in listedOn:
                monthNum = 11
            elif "Oct" in listedOn:
                monthNum = 10
            elif "Sep" in listedOn:
                monthNum = 9
            elif "Jun" in listedOn:
                monthNum = 6
            elif "Apr" in listedOn:
                monthNum = 4
            elif "Jan" in listedOn:
                monthNum = 1
            elif "Feb" in listedOn:
                monthNum = 2
            elif "Mar" in listedOn:
                monthNum = 3
            elif "May" in listedOn:
                monthNum = 5
            elif "Aug" in listedOn:
                monthNum = 8
            elif "Dec" in listedOn:
                monthNum = 12
            elif "Jul" in listedOn:
                monthNum = 7
            
            if "th" in listedOn:
                day = listedOn.split("th")[0]
            elif "nd" in listedOn:
                day = listedOn.split("nd")[0]
            elif "st" in listedOn:
                day = listedOn.split("st")[0]
            elif "rd" in listedOn:
                day = listedOn.split("rd")[0]
                
            year = "2017" 
            
            listedOn = day + "/" + str(monthNum) + "/" + year
            #print("listedOn - after:", listedOn)
                    
            """ get ID """
            #idPro = blockProperty["id"].split("ng_")[1]
            idPro = blockProperty.parent["data-listing-id"]
                    
            """ get Price"""
            pricePro_2 = []
            pricePro = blockProperty.find_all("", {"class":"listing-results-price"})[0].text.replace(" ", "")
            pricePro = pricePro.split("\n")
            for item in pricePro:
                if item:
                    pricePro_2.append(str(item))
            pricePro = pricePro_2[0].replace("£", "")
                    
            """ get Bedrooms """
            bedrooms = 0
            try:
                blockForAttrs = blockProperty.find("h3").find_all("span")
                for x in range(0, len(blockForAttrs), 1):
                    try:
                        if "num-beds" in blockForAttrs[x]["class"]:
                            bedrooms = float(blockForAttrs[x].text)
                    except:
                        print("No beds")
            except:
                bedrooms = 0
                        
            """ get Bathrooms """
            bathrooms = 0
            try:
                blockForAttrs = blockProperty.find("h3").find_all("span")
                bathrooms = 0
                for x in range(0, len(blockForAttrs), 1):
                    try:
                        if "num-baths" in blockForAttrs[x]["class"]:
                            bathrooms = float(blockForAttrs[x].text)
                    except:
                        print("No bath")
            except:
                bathrooms = 0
                        
            """ get Reception rooms """
            receptions = 0
            try:
                blockForAttrs = blockProperty.find("h3").find_all("span")
                receptions = 0
                for x in range(0, len(blockForAttrs), 1):
                    try:
                        if "num-reception" in blockForAttrs[x]["class"]:
                            receptions = float(blockForAttrs[x].text)
                    except:
                        print("No bath")
            except:
                receptions = 0
                    
            ppM = pricePro.split("pcm")[0].replace(",", "")
            #ppW = pricePro.split("pcm")[1].replace(",","")
            #ppW = re.findall(r"\d+", ppW)[0]
            #print("Price per Month:", ppM)
            
            """ get description """
            try:
                descriptionText = blockProperty.find_all("p", {"itemprop":"description"})[0].text.split("\n")
                for item in descriptionText:
                    if len(item) < 1:
                        descriptionText.remove(item)
                descriptionText = descriptionText[0]
            except:
                print("No description")
                descriptionText = "No description"
                
            if len(descriptionText) > 1:
                whiteIndex = 0
                z = 0
                while descriptionText[z] == " ":
                    whiteIndex = whiteIndex + 1
                    z = z + 1
                descriptionText = descriptionText[z:]
                
            """ get nearby station : working """
            nearbyStation = ""
            station_here = blockProperty.find("div", {"class":"nearby_stations_schools"}).text.split("\n")
            #station_here = ''.join(station_here).split()
            z = 0
            while len(station_here[z]) < 2:
                z = z + 1
            nearbyStation = station_here[z]
            
            """ Trying to get Lot Lan"""
            #geoBlock = blockProperty.find_all("div", {"itemprop":"geo"})
            geo_lat = blockProperty.find(itemprop = "latitude").get("content")
            geo_lon = blockProperty.find(itemprop = "longitude").get("content")
            #print("geo_lat", geo_lat)
            #print("geo_lon", geo_lon)
            print("----------------")
                
            """ get Type : working """
            finalType = ""
            typeProp = blockProperty.find("h2").find_all("a")[0].text
            if "flat".upper() in typeProp.upper():
                finalType = "flat"
            elif "semi-deteached".upper() in typeProp.upper():
                    finalType = "semi-deteached house"
            elif "property".upper() in typeProp.upper():
                finalType = "apartment"
            elif "studio".upper() in typeProp.upper():
                finalType = "studio"
            elif "maisonette".upper() in typeProp.upper():
                finalType = "maisonette"
            elif "room".upper() in typeProp.upper():
                finalType = "room"
            elif "house".upper() in typeProp.upper():
                finalType = "house"
            elif "shared accommodation".upper() in typeProp.upper():
                finalType = "shared accomodation"
            #print(finalType, type(finalType))
                    
            """ get link """
            linkTo = str(blockProperty.find("a")).split('"/')[1].split('"')[0]
            linkTo = "https://www.zoopla.co.uk/" + linkTo
                    
            """ get full descriptiom """
            descriptionFull = getFullDescription(linkTo)
            
            """ get Floor plan """
            #FloorPlanURL = getFloorPlan(linkTo)
                    
            """ Save features to Database """
            featuresOfProperties["ID"] = idPro
            featuresOfProperties["DESCRIPTION"] = descriptionText
            featuresOfProperties["STATION"] = nearbyStation
            featuresOfProperties["ADDED"] = listedOn
            featuresOfProperties["LOCATION"] = address
            featuresOfProperties["PRICE PCM (£)"] = ppM
            featuresOfProperties["TYPE"] = finalType
            featuresOfProperties["BATHROOMS"] = bathrooms
            featuresOfProperties["BEDROOMS"] = bedrooms
            featuresOfProperties["RECEPTIONS"] = receptions
            featuresOfProperties["LINK"] = linkTo
            featuresOfProperties["FULL DESCRIPTION"] = descriptionFull
            featuresOfProperties["LON"] = geo_lon
            featuresOfProperties["LAT"] = geo_lat
                
            properties.append(dict(featuresOfProperties))
            #print("---------------------")
    return properties
""" Get Data : End"""

""" Update EXCEL file : start """
def updateExcel(dataset):
    import openpyxl, os, datetime
    from datetime import datetime
    from pandas import ExcelWriter as ewriter
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Fill
    
    #directory = "C:\\Users\\Vytautas.Bielinskas\\Desktop\\Python\\01 Web Scraping\\Zoopla"
    #os.chdir(directory)
    #os.getcwd()
    
    filename = "DataSaves - Data rent.xlsx"
    wb = openpyxl.load_workbook(filename, data_only = False)    
    
    #print(wb.get_sheet_names)       
    
    sheet = wb.get_sheet_by_name("Data Rents")
    
    rowForSearching = 10
    while len(str(sheet.cell(row = rowForSearching, column = 2).value)) > 4:
        rowForSearching = rowForSearching + 1
        #print(rowForSearching, str(sheet.cell(row = rowForSearching, column = 2).value), len(str(sheet.cell(row = rowForSearching, column = 2).value)))
        
    #print("rowForSearching: ", rowForSearching)
    
    """ Define some Excel cell styles and formatting """
    fillDefault = PatternFill(fill_type = None,
                              start_color = "FFFFFFFF",
                              end_color = "FF000000")
    
    HyperlinkBlue = Font(color = "0563c1",
                         underline = "single")
    
    rightAligment = Alignment(horizontal = "right")
    centerAligment = Alignment(horizontal = "center")
    
    """ Starting writing data to Excel file """
    id_column = dataset.columns.get_loc("ID")                        # get index of ID column
    id_station = dataset.columns.get_loc("STATION")                  # get index of STATION column
    id_location = dataset.columns.get_loc("LOCATION")                # get index of LOCATION column
    id_bedrooms = dataset.columns.get_loc("BEDROOMS")                # get index of BEDROOMS column
    id_type = dataset.columns.get_loc("TYPE")                        # get index of TYPE column
    id_pricePCM = dataset.columns.get_loc("PRICE PCM (£)")           # get index of PCM column
    id_description = dataset.columns.get_loc("DESCRIPTION")          # get index of DESCRIPTION column
    id_link = dataset.columns.get_loc("LINK")                        # get index of LINK column
    id_added = dataset.columns.get_loc("ADDED")                      # get index of ADDED column
    
    dataset_index = 0
    for i in range(rowForSearching, rowForSearching + len(dataset), 1):
        #print(i)
        """ writing MONTH column """
        today = datetime.now().strftime('%d')
        month = datetime.now().strftime('%m')
        year = "17"
        if month == "11":
            monthW = "Nov"
        dateNow = today + "-" + monthW + "-" + year
        sheet.cell(row = i, column = 2).value = dateNow
        sheet.cell(row = i, column = 2).alignment = rightAligment
        
        """ writing ID column """
        sheet.cell(row = i, column = 3).value = dataset.iat[dataset_index, id_column]
        
        """ writing STATION column """
        sheet.cell(row = i, column = 4).value = dataset.iat[dataset_index, id_station].upper()
        
        """ writing LOCATION column """
        sheet.cell(row = i, column = 5).value = dataset.iat[dataset_index, id_location]
        
        """ writing BEDROOMS column """
        sheet.cell(row = i, column = 6).value = dataset.iat[dataset_index, id_bedrooms]
        sheet.cell(row = i, column = 6).alignment = centerAligment
        
        """ writing TYPE column """
        sheet.cell(row = i, column = 7).value = dataset.iat[dataset_index, id_type]
        sheet.cell(row = i, column = 7).alignment = centerAligment
        
        """ writing PRICE PCM column """
        sheet.cell(row = i, column = 8).value = dataset.iat[dataset_index, id_pricePCM]
        sheet.cell(row = i, column = 8).alignment = centerAligment
        
        """ writing DESCRIPTION column """
        sheet.cell(row = i, column = 10).value = dataset.iat[dataset_index, id_description]
        
        """ writing LINK column """
        FullLink = '=HYPERLINK("' + dataset.iat[dataset_index, id_link] + '","' + "WEBPAGE" + '")'
        sheet.cell(row = i, column = 11).value = FullLink
        sheet.cell(row = i, column = 11).font = HyperlinkBlue
        sheet.cell(row = i, column = 11).alignment = centerAligment
        
        """ writing ADDED column """
        sheet.cell(row = i, column = 14).value = dataset.iat[dataset_index, id_added]
        sheet.cell(row = i, column = 14).fill = fillDefault
        
        dataset_index = dataset_index + 1
        
    for i in range(rowForSearching, sheet.max_row, 1)    :
        for j in range(1, sheet.max_column, 1):
            sheet.cell(row = i, column = j).fill = fillDefault
    
    """ SAVE DATA TO EXCEL """
    import datetime
    from datetime import date 
    timeNow = str(datetime.datetime.now()).replace(":","-")[:-10]
    wb.save("DataSaved-FINAL - RENT - " + timeNow + ".xlsx")
        
    return None

""" Update EXCEL file : end """

""" Save data to file : start """
def writeToFile(DF):
    import time, datetime, os
    from datetime import date 
    
    os.chdir("C:\\Users\\Vytautas.Bielinskas\\Desktop\\Python\\01 Web Scraping\\Zoopla")
    timeNow = str(datetime.datetime.now()).replace(":","-")[:-10]
    DF.to_csv("Zoopla.com - RENTS data.csv " + timeNow + ".csv")
    return None
""" Save data to file : end """
        
numberOfPages = 7
URLs = generateURLs(numberOfPages)
FullData = getData(URLs)
DF = pd.DataFrame(FullData)
DF = DF[["ID", "STATION", "LOCATION", "BEDROOMS", "BATHROOMS", "RECEPTIONS", "TYPE", "PRICE PCM (£)", "DESCRIPTION", "LINK", "ADDED", "FULL DESCRIPTION", "LAT", "LON"]]

DataBase = DF
writeToFile(DataBase)
#updateExcel(DataBase)